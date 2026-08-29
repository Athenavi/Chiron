"""Media tools 注册到本地工具注册表。

实现对标 Go `internal/tools/media.go` 注册的两个工具：
- media_create：创建媒体资产（文本/CSV/代码等）
- image_generate：生成图片（AI 生成失败时降级为 SVG 占位图）

默认使用本地 MediaStore；后续可接入 S3 + DB。
"""

from __future__ import annotations

import base64
import logging
import os
import re
from datetime import datetime
from typing import Any

import httpx

from app.media.store import create_store
from app.tools.registry import registry
from app.tools.ssrf import assert_safe_url, fetch_url_safe

logger = logging.getLogger(__name__)
_store = create_store()

# 文件下载大小限制：100MB
MAX_DOWNLOAD_SIZE = 100 * 1024 * 1024


def _sanitize_filename(prompt: str) -> str:
    name = re.sub(r"[^a-zA-Z0-9 _\-]", "", prompt).strip()
    name = re.sub(r"\s+", "_", name)
    return name[:48] or "image"


# ── media_create ──────────────────────────────────────────────
async def media_create(
    name: str,
    content: str,
    type: str = "text",
    category: str = "generated",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    if not name or not content:
        return {"error": "name and content are required"}
    data = content.encode("utf-8")
    asset = _store.write(
        name=name, content=data, asset_type=type, category=category, tags=tags or []
    )
    return {
        "output": f"Media asset '{name}' created ({asset.size} bytes)",
        "id": asset.id,
        "name": asset.name,
        "type": asset.type,
        "category": asset.category,
        "file_url": asset.file_url,
        "size": asset.size,
    }


# ── image_generate ────────────────────────────────────────────
async def image_generate(
    prompt: str = "Generated Image",
    width: int = 800,
    height: int = 600,
    category: str = "generated",
) -> dict[str, Any]:
    """真实生成图片；未配置 provider 时明确报错（S 修复：移除假的 SVG 占位，
    不再虚报生成成功）。配置 IMAGE_GEN_API_URL(+IMAGE_GEN_API_KEY) 接入生成服务。"""
    width = max(64, min(width, 4096))
    height = max(64, min(height, 4096))

    api_url = os.getenv("IMAGE_GEN_API_URL", "").strip()
    api_key = os.getenv("IMAGE_GEN_API_KEY", "").strip()
    if not api_url:
        return {
            "error": "image generation not available: IMAGE_GEN_API_URL not configured"
        }

    try:
        from app.config import settings
        headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
        async with httpx.AsyncClient(timeout=settings.http_timeout_long) as client:
            resp = await client.post(
                api_url,
                json={"prompt": prompt, "width": width, "height": height},
                headers=headers,
            )
            resp.raise_for_status()
            data = resp.content
    except Exception as e:  # noqa: BLE001 — 生成失败如实返回，不伪造成功
        return {"error": f"image generation failed: {e}"}

    name = _sanitize_filename(prompt) + ".png"
    asset = _store.write(
        name=name,
        content=data,
        asset_type="image",
        category=category,
        fmt="image/png",
        width=width,
        height=height,
    )

    return {
        "output": f"Image generated: {name} ({asset.size} bytes)",
        "id": asset.id,
        "name": name,
        "type": "image",
        "format": "image/png",
        "width": width,
        "height": height,
        "category": category,
        "file_url": asset.file_url,
        "size": asset.size,
    }


# ── vision_analyze: 图片理解 ─────────────────────────────────
async def vision_analyze(
    image_url: str,
    prompt: str = "请详细描述这张图片的内容。",
    detail: str = "auto",
) -> dict[str, Any]:
    """使用多模态模型（GPT-4V / Claude-3 Vision）分析图片内容。

    Args:
        image_url: 图片 URL（支持 http/https 或 data:image 格式）
        prompt: 分析提示词
        detail: 细节级别 "auto" | "low" | "high"

    Returns:
        包含分析结果和所用模型的字典
    """
    from app.media.analyzer import analyze_image

    result = await analyze_image(image_url, prompt=prompt)
    if result.get("success"):
        return result
    # analyze_image 失败时降级：用 llm 客户端直接调用
    try:
        from app.config import settings

        # SSRF 防护：跳过 data: URL
        if not image_url.startswith("data:"):
            assert_safe_url(image_url)

        async with httpx.AsyncClient(timeout=settings.http_timeout_web) as client:
            resp = await client.head(image_url, follow_redirects=False)
            content_type = resp.headers.get("content-type", "")
            if not content_type.startswith("image/"):
                return {"error": f"URL does not point to an image: {content_type}"}

        from app.llm.client import get_llm_client

        llm = get_llm_client()
        model = os.getenv("VISION_MODEL", "gpt-4o")
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {"url": image_url, "detail": detail},
                    },
                ],
            }
        ]
        response = await llm.chat(messages, model=model)
        return {
            "success": True,
            "analysis": (
                response.content if hasattr(response, "content") else str(response)
            ),
            "model": model,
            "analyzed_at": datetime.utcnow().isoformat(),
        }
    except Exception as e:
        logger.error(f"Vision analysis failed: {e}")
        return {"success": False, "error": str(e)}


# ── speech_to_text: 语音转文字 ────────────────────────────────
async def speech_to_text(
    audio_url: str,
    language: str = "zh",
    prompt: str = "",
) -> dict[str, Any]:
    """将音频文件转为文字（Whisper API）。

    Args:
        audio_url: 音频文件 URL（支持 http/https 和 data:audio 格式）
        language: 音频语言代码（zh/en/ja 等）
        prompt: 可选的提示词，帮助模型理解上下文

    Returns:
        包含识别文本的字典
    """
    api_key = os.getenv("OPENAI_API_KEY", "")
    base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
    if not api_key:
        # 尝试从 settings 获取
        from app.config import settings

        api_key = settings.openai_api_key or settings.llm_api_key or ""
    if not api_key:
        return {
            "error": "speech_to_text not available: no OpenAI API key configured (set OPENAI_API_KEY)"
        }

    # SSRF 防护
    assert_safe_url(audio_url)

    try:
        from app.config import settings as s

        async with httpx.AsyncClient(timeout=s.http_timeout_long, follow_redirects=False) as client:
            resp = await fetch_url_safe(client, audio_url)
            resp.raise_for_status()
            audio_data = resp.content
            if len(audio_data) > MAX_DOWNLOAD_SIZE:
                return {"error": f"audio file too large: {len(audio_data)} bytes (max {MAX_DOWNLOAD_SIZE})"}

        # 调用 Whisper API
        whisper_url = f"{base_url.rstrip('/')}/audio/transcriptions"
        assert_safe_url(whisper_url)
        files = {"file": ("audio.wav", audio_data, resp.headers.get("content-type", "audio/wav"))}
        data = {"model": "whisper-1", "language": language, "response_format": "json"}
        if prompt:
            data["prompt"] = prompt

        async with httpx.AsyncClient(timeout=s.http_timeout_long) as client:
            whisper_resp = await client.post(
                whisper_url,
                headers={"Authorization": f"Bearer {api_key}"},
                files=files,
                data=data,
            )
            whisper_resp.raise_for_status()
            result = whisper_resp.json()

        return {
            "success": True,
            "text": result.get("text", ""),
            "language": language,
            "duration_seconds": result.get("duration", 0),
        }
    except Exception as e:
        logger.error(f"Speech to text failed: {e}")
        return {"error": f"speech_to_text failed: {e}"}


# ── text_to_speech: 文字转语音 ────────────────────────────────
async def text_to_speech(
    text: str,
    voice: str = "alloy",
    model: str = "tts-1",
    speed: float = 1.0,
    output_format: str = "mp3",
) -> dict[str, Any]:
    """将文字转为语音音频（OpenAI TTS API）。

    Args:
        text: 要转为语音的文字
        voice: 音色（alloy/echo/fable/onyx/nova/shimmer）
        model: TTS 模型（tts-1 / tts-1-hd）
        speed: 语速（0.25 ~ 4.0）
        output_format: 输出格式（mp3 / opus / aac / flac / wav）

    Returns:
        包含音频 data URL 的字典
    """
    api_key = os.getenv("OPENAI_API_KEY", "")
    base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
    if not api_key:
        from app.config import settings

        api_key = settings.openai_api_key or settings.llm_api_key or ""
    if not api_key:
        return {
            "error": "text_to_speech not available: no OpenAI API key configured (set OPENAI_API_KEY)"
        }

    from app.config import settings as s

    speed = max(0.25, min(speed, 4.0))
    valid_voices = {"alloy", "echo", "fable", "onyx", "nova", "shimmer"}
    if voice not in valid_voices:
        voice = "alloy"
    valid_formats = {"mp3", "opus", "aac", "flac", "wav"}
    if output_format not in valid_formats:
        output_format = "mp3"

    try:
        # SSRF 防护：校验 OPENAI_BASE_URL 不指向内网
        tts_url = f"{base_url.rstrip('/')}/audio/speech"
        assert_safe_url(tts_url)
        async with httpx.AsyncClient(timeout=s.http_timeout_long) as client:
            resp = await client.post(
                tts_url,
                headers={"Authorization": f"Bearer {api_key}"},
                json={
                    "model": model,
                    "input": text,
                    "voice": voice,
                    "speed": speed,
                    "response_format": output_format,
                },
            )
            resp.raise_for_status()
            audio_data = resp.content

        mime_map = {"mp3": "audio/mpeg", "opus": "audio/opus", "aac": "audio/aac", "flac": "audio/flac", "wav": "audio/wav"}
        data_url = f"data:{mime_map.get(output_format, 'audio/mpeg')};base64,{base64.b64encode(audio_data).decode('ascii')}"

        return {
            "success": True,
            "data_url": data_url,
            "format": output_format,
            "bytes": len(audio_data),
            "voice": voice,
            "text_length": len(text),
        }
    except Exception as e:
        logger.error(f"Text to speech failed: {e}")
        return {"error": f"text_to_speech failed: {e}"}


# ── file_analyzer: 文件分析 ───────────────────────────────────
async def file_analyzer(
    file_url: str,
    analysis_type: str = "auto",
    custom_prompt: str = "",
) -> dict[str, Any]:
    """分析文件内容（PDF/CSV/Excel/图片/文本等）。

    Args:
        file_url: 文件 URL
        analysis_type: 分析类型（auto/pdf/csv/image/text）
        custom_prompt: 自定义分析提示词

    Returns:
        包含分析结果的字典
    """
    # SSRF 防护
    assert_safe_url(file_url)

    from app.config import settings as s

    try:
        # 下载文件
        async with httpx.AsyncClient(timeout=s.http_timeout_long, follow_redirects=False) as client:
            resp = await fetch_url_safe(client, file_url)
            resp.raise_for_status()
            file_data = resp.content
            content_type = resp.headers.get("content-type", "")

        # 文件大小限制（100MB）
        if len(file_data) > MAX_DOWNLOAD_SIZE:
            return {"error": f"file too large: {len(file_data)} bytes (max {MAX_DOWNLOAD_SIZE})"}

        # 根据类型决定分析策略
        is_image = content_type.startswith("image/") or analysis_type == "image"
        is_pdf = "pdf" in content_type or analysis_type == "pdf" or file_url.lower().endswith(".pdf")
        is_csv = "csv" in content_type or analysis_type == "csv" or file_url.lower().endswith(".csv")
        is_excel = "spreadsheet" in content_type or analysis_type in ("excel", "xlsx") or file_url.lower().endswith((".xlsx", ".xls"))

        if is_image and analysis_type != "text":
            b64 = base64.b64encode(file_data).decode("ascii")
            data_url = f"data:{content_type};base64,{b64}"
            return await vision_analyze(
                image_url=data_url,
                prompt=custom_prompt or "请详细分析这张图片的内容，包括主要对象、场景、文字（如有）。",
            )

        # 文本类文件 → 提取文本后用 LLM 分析
        text = ""
        if is_pdf:
            import pymupdf

            doc = pymupdf.open(stream=file_data, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
        elif is_csv:
            text = file_data.decode("utf-8", errors="replace")
            lines = text.splitlines()
            if len(lines) > 50:
                text = "\n".join(lines[:50]) + f"\n\n... (共 {len(lines)} 行，仅显示前 50 行)"
        elif is_excel:
            import openpyxl
            import io

            wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                preview = rows[:20]
                parts.append(f"=== Sheet: {sheet_name} ({len(rows)} rows) ===")
                for row in preview:
                    parts.append(", ".join(str(c) if c is not None else "" for c in row))
            text = "\n".join(parts)
            wb.close()
        else:
            # 文本文件
            text = file_data.decode("utf-8", errors="replace")[:50000]

        if not text.strip():
            return {"error": "无法提取文件内容，不支持的格式或空文件"}

        # 用 LLM 分析
        from app.llm.client import get_llm_client

        llm = get_llm_client()
        prompt_text = custom_prompt or f"""请分析以下文件内容：

{text[:30000]}

请提供：
1. 文件类型和概览
2. 主要内容摘要
3. 关键数据/发现
4. 可能的用途或建议"""
        response = await llm.chat([{"role": "user", "content": prompt_text}])
        return {
            "success": True,
            "analysis": response.content if hasattr(response, "content") else str(response),
            "file_type": content_type,
            "text_length": len(text),
            "analyzed_at": datetime.utcnow().isoformat(),
        }
    except Exception as e:
        logger.error(f"File analysis failed: {e}")
        return {"error": f"file_analyzer failed: {e}"}


# ── 注册 ──────────────────────────────────────────────────────
registry.register(
    name="media_create",
    description="Create a media asset (text, CSV, code, etc.).",
    parameters={
        "type": "object",
        "properties": {
            "name": {"type": "string"},
            "content": {"type": "string"},
            "type": {"type": "string", "default": "text"},
            "category": {"type": "string", "default": "generated"},
            "tags": {"type": "array", "items": {"type": "string"}, "default": []},
        },
        "required": ["name", "content"],
    },
    handler=media_create,
)

registry.register(
    name="image_generate",
    description="Generate an image from a text prompt (requires IMAGE_GEN_API_URL; fails loudly when unconfigured).",
    parameters={
        "type": "object",
        "properties": {
            "prompt": {"type": "string", "default": "Generated Image"},
            "width": {"type": "integer", "default": 800},
            "height": {"type": "integer", "default": 600},
            "category": {"type": "string", "default": "generated"},
        },
    },
    handler=image_generate,
)

registry.register(
    name="vision_analyze",
    description="Analyze an image using a vision-capable model (GPT-4V / Claude-3 Vision). Provide an image URL and optional prompt.",
    parameters={
        "type": "object",
        "properties": {
            "image_url": {"type": "string", "description": "Image URL (http/https or data:image)"},
            "prompt": {"type": "string", "default": "请详细描述这张图片的内容。"},
            "detail": {"type": "string", "enum": ["auto", "low", "high"], "default": "auto"},
        },
        "required": ["image_url"],
    },
    handler=vision_analyze,
)

registry.register(
    name="speech_to_text",
    description="Transcribe audio to text using Whisper API. Provide an audio URL and optional language code.",
    parameters={
        "type": "object",
        "properties": {
            "audio_url": {"type": "string", "description": "Audio file URL (http/https)"},
            "language": {"type": "string", "default": "zh", "description": "Audio language code (zh/en/ja)"},
            "prompt": {"type": "string", "default": "", "description": "Optional prompt to guide the model"},
        },
        "required": ["audio_url"],
    },
    handler=speech_to_text,
)

registry.register(
    name="text_to_speech",
    description="Convert text to speech audio using OpenAI TTS API. Returns a data URL with the audio content.",
    parameters={
        "type": "object",
        "properties": {
            "text": {"type": "string", "description": "Text to convert to speech"},
            "voice": {"type": "string", "enum": ["alloy", "echo", "fable", "onyx", "nova", "shimmer"], "default": "alloy"},
            "model": {"type": "string", "default": "tts-1"},
            "speed": {"type": "number", "default": 1.0, "description": "Speech speed (0.25-4.0)"},
            "output_format": {"type": "string", "enum": ["mp3", "opus", "aac", "flac", "wav"], "default": "mp3"},
        },
        "required": ["text"],
    },
    handler=text_to_speech,
)

registry.register(
    name="file_analyzer",
    description="Analyze a file (PDF, CSV, Excel, image, text) and extract its content and insights.",
    parameters={
        "type": "object",
        "properties": {
            "file_url": {"type": "string", "description": "File URL to analyze"},
            "analysis_type": {
                "type": "string",
                "enum": ["auto", "pdf", "csv", "image", "text", "excel"],
                "default": "auto",
                "description": "Force a specific analysis type",
            },
            "custom_prompt": {"type": "string", "default": "", "description": "Custom analysis prompt"},
        },
        "required": ["file_url"],
    },
    handler=file_analyzer,
)
